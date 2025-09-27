from datetime import datetime
import os
import io # Para gestionar el archivo en memoria
from django.conf import settings
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.styles import Font

from apps.cargo.models.cargo import Cargo
from apps.constancia.models.anio import Anio
from apps.constancia.views.constancia import generar_constancia, get_cargo

FORMATO_DECIMAL = "0.00"
def obtener_valores_para_excel(data_list):
    """Convierte la lista de contexto de Django a una lista de listas de valores simples."""
    valores_excel = []

    for row in data_list:
        # Columna A (código), Columna B (concepto)
        codcon = getattr(row[0], 'codcon', '')
        descon = getattr(row[0], 'descon', '')

        fila_excel = [codcon, descon]

        # Columnas C a N (Enero a Diciembre)
        # Los montos están en los índices 1 a 12
        for i in range(1, 13):
            monto_obj = row[i]
            monto = getattr(monto_obj, 'monto', 0) if isinstance(monto_obj, object) else 0
            fila_excel.append(monto or 0) # Si es None o 0, usa 0

        # Columna O (Total) - El total está en el índice 13
        total = row[13] if len(row) > 13 else 0
        fila_excel.append(total or 0) # Si es None o 0, usa 0

        valores_excel.append(fila_excel)

    return valores_excel

def rellenar_seccion_dinamica(sheet, data_list, start_row):
    """Escribe filas de datos a partir de una fila de inicio dada."""
    for fila_idx, fila_datos in enumerate(data_list):
        fila_para_escribir = start_row + fila_idx

        for col_offset, valor in enumerate(fila_datos):
            columna_actual = 1 + col_offset # Columna 1 = A

            celda = sheet.cell(row=fila_para_escribir, column=columna_actual, value=valor)
            celda.font = Font(size=8)
            # Aplicar formato de número si es Columna C o posterior (columna_actual >= 3)
            if columna_actual >= 3:
                # El valor ya debe ser float gracias a obtener_valores_para_excel()
                celda.number_format = FORMATO_DECIMAL

    # Devuelve la fila donde debe comenzar la siguiente sección (una después del último dato)
    return start_row + len(data_list)

def aplicar_total_y_estilo(sheet, total_list, fila_total):
    """Escribe la fila de totales y aplica el estilo."""

    # El primer elemento del total (el texto 'TOTAL INGRESOS') va en la Columna B (índice 2)
    sheet.cell(row=fila_total, column=2, value=total_list[0]).font = Font(size=8,bold=True)

    # Los montos van desde la Columna C (índice 3) hasta la O (índice 15)
    for col_offset, valor in enumerate(total_list[1:]):
        columna_actual = 3 + col_offset # Empieza en C (3)
        celda = sheet.cell(row=fila_total, column=columna_actual, value=valor)
        celda.font = Font(size=8,bold=True)
        # Aplicar formato decimal
        celda.number_format = FORMATO_DECIMAL
    # Devuelve la fila donde debe comenzar la siguiente sección (una después del total)
    return fila_total + 1

# Asegúrate de que esta función está en tu 'views.py'
def print_constancia_excel(request):

    pk = request.GET.get('pk', 1)

    ruta_relativa = os.path.join('..', 'apps', 'constancia', 'resources', 'plantilla_constancia.xlsx')
    ruta_plantilla = os.path.join(settings.BASE_DIR, ruta_relativa)
    constancia = generar_constancia(request, pk)

    # 2. Cargar la plantilla con openpyxl
    try:
        workbook = load_workbook(ruta_plantilla)
    except FileNotFoundError:
        return HttpResponse("Error: Plantilla de Excel no encontrada.", status=500)


    # 3. Escribir datos en la plantilla
    sheet = workbook.active  # O workbook['Nombre de la Hoja']
    # 🚨 Variable clave: Fila actual donde comenzaremos a escribir los datos de la tabla.
    anio_id = request.session["anio"]
    anio_bd = request.session["anio_bd"]
    anio = Anio.objects.get(pk=anio_id)
    sheet['A11'] = 'AÑO ' + str(anio.anio)
    sheet['C14'] = constancia['constancia'].plaza
    sheet['D14'] = constancia['trabajador'].nombre_completo
    sheet['C15'] = constancia['constancia'].nivel
    # sheet['C16'] = dependencia
    cargo = get_cargo(anio_bd,constancia['trabajador'],anio.id)
    sheet['J14'] = cargo
    sheet['J15'] = constancia['trabajador'].dni
    sheet['J16'] = datetime.now().date()

    fila_actual = 19
    # Define el estilo de negrita para los totales
    ingresos_excel = obtener_valores_para_excel(constancia['monto_ingreso_list'])
    fila_actual = rellenar_seccion_dinamica(sheet, ingresos_excel, fila_actual)
    # Total de Ingresos
    fila_actual = aplicar_total_y_estilo(sheet, constancia['monto_ingreso_list_total'], fila_actual)
    fila_actual = fila_actual +1

    descuentos_excel = obtener_valores_para_excel(constancia['monto_descuento_list'])
    fila_actual = rellenar_seccion_dinamica(sheet, descuentos_excel, fila_actual)
    # Total de Descuentos
    fila_actual = aplicar_total_y_estilo(sheet, constancia['monto_descuento_list_total'], fila_actual)
    fila_actual = fila_actual +1

    aportaciones_excel = obtener_valores_para_excel(constancia['monto_aportaciones_list'])
    fila_actual = rellenar_seccion_dinamica(sheet, aportaciones_excel, fila_actual)
    # Total de aportaciones
    fila_actual = aplicar_total_y_estilo(sheet, constancia['monto_aportaciones_list_total'], fila_actual)
    fila_actual = fila_actual +1

    # Total de liquido
    fila_actual = aplicar_total_y_estilo(sheet, constancia['monto_liquido_list_total'], fila_actual)

    # footer
    fila_actual = fila_actual + 4
    f1 = sheet.cell(row=fila_actual, column=1,value='Fuente: Planilla Unica de Pago de Haberes')
    f1.font = Font(size=8)
    f2 = sheet.cell(row=fila_actual, column=13,value='Realizado por: ')
    f2.font = Font(size=8)
    fila_actual = fila_actual + 1
    f3 = sheet.cell(row=fila_actual, column=1,value='Los funcionarios que firman al pie del presente declaran bajo responsabilidad que la liquidacion presente es veraz y está respaldada por las planillas de pago.')
    f3.font = Font(size=8)
    # Crea un buffer (almacenamiento temporal) en memoria RAM
    excel_buffer = io.BytesIO()

    # Guarda el 'workbook' modificado en el buffer
    workbook.save(excel_buffer)
    excel_buffer.seek(0) # Vuelve al inicio del buffer

    # 5. CREACIÓN DE LA RESPUESTA HTTP
    response = HttpResponse(
        excel_buffer.read(), # Lee el contenido del buffer
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # 6. CONFIGURAR ENCABEZADOS DE DESCARGA
    nombre_archivo = 'ReporteHaberes_'+str(anio.anio)+'_'+constancia['trabajador'].apellido_paterno+'_'+datetime.now().date().strftime('%d-%m-%Y')+'.xlsx'

    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

    return response

